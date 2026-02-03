#!/usr/bin/env python3
"""
MAHLE ROM Data Updater
======================
This script extracts data from the MAHLE ROM Excel file and updates the
rom-data.json file that powers the web presentation.

Usage:
    python update_rom_data.py [path_to_rom_excel]

If no path is provided, it looks for the ROM file in the parent directory.

Requirements:
    pip install openpyxl
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


def find_rom_file():
    """Look for the ROM Excel file in common locations."""
    possible_paths = [
        "../MAHLE_ROM_2026_Populated_v1.xlsx",
        "MAHLE_ROM_2026_Populated_v1.xlsx",
        "../MAHLE_ROM*.xlsx",
    ]

    for path in possible_paths:
        if os.path.exists(path):
            return path

    # Try glob pattern
    import glob
    matches = glob.glob("../*ROM*.xlsx")
    if matches:
        return matches[0]

    return None


def extract_rom_data(excel_path):
    """Extract key data from the ROM Excel file."""
    print(f"Reading ROM data from: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Initialize data structure
    rom_data = {
        "meta": {
            "lastUpdated": datetime.now().strftime("%Y-%m-%d"),
            "version": "1.0",
            "client": "MAHLE",
            "project": "AI Finance Transformation",
            "sourceFile": os.path.basename(excel_path)
        },
        "summary": {},
        "team": [],
        "phases": {},
        "totals": {},
        "businessCase": {},
        "valueProjections": {},
        "timeline": {},
        "useCases": [],
        "staffing": []
    }

    # Extract from "Drivers & Estimates" sheet
    if "Drivers & Estimates" in wb.sheetnames:
        sheet = wb["Drivers & Estimates"]
        rom_data["summary"] = {
            "investmentLow": 260000,
            "investmentAvg": 300000,
            "investmentHigh": 370000,
            "weeksLow": 10,
            "weeksAvg": 12,
            "weeksHigh": 14,
            "sprintsLow": 5,
            "sprintsAvg": 6,
            "sprintsHigh": 7,
            "targetVelocity": 60,
            "blendedRate": 108.38
        }

        # Try to extract actual values from sheet
        for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
            if row[0] and "velocity" in str(row[0]).lower():
                try:
                    rom_data["summary"]["targetVelocity"] = float(row[5]) if row[5] else 60
                except:
                    pass

    # Extract from "Points" sheet
    if "Points" in wb.sheetnames:
        sheet = wb["Points"]
        phases = {}
        for row in sheet.iter_rows(min_row=2, max_row=10, values_only=True):
            if row[0] and row[1] is not None:
                phase_name = str(row[0]).replace(" ", "").replace("&", "")
                phases[phase_name] = {
                    "pointsLow": int(row[1]) if row[1] else 0,
                    "pointsHigh": int(row[2]) if row[2] else 0,
                    "features": int(row[3]) if row[3] else 0
                }
        if phases:
            rom_data["phases"] = phases

    # Extract from "Fees & costs" sheet
    if "Fees & costs" in wb.sheetnames:
        sheet = wb["Fees & costs"]
        staffing = []

        for row in sheet.iter_rows(min_row=10, max_row=25, values_only=True):
            if row[0] and row[1] and "total" not in str(row[0]).lower():
                role = str(row[0]).strip()
                if role and role not in ["", "Add rows as necessary"]:
                    staffing.append({
                        "role": role,
                        "fte": float(row[1]) if row[1] else 0,
                        "costRate": float(row[2]) if row[2] else 0
                    })

        if staffing:
            rom_data["staffing"] = staffing

    # Business case data (these would need to be updated manually or from another source)
    rom_data["businessCase"] = {
        "currentDSO": 60,
        "targetDSOReduction": 10,
        "receivables": 70000000,
        "dailySalesRate": 1166667,
        "workingCapitalFreed": 11666670,
        "costOfCapital": 0.035,
        "annualCapitalSavings": 408333,
        "vendorInquiryFTEs": 15,
        "deflectionRateTarget": 0.65,
        "apTimeSavingsTarget": 0.25,
        "invoiceCycleCurrentDays": 5,
        "invoiceCycleTargetDays": 2,
        "aiAccuracyTarget": 0.80
    }

    # Value projections
    rom_data["valueProjections"] = {
        "workingCapitalValue": 400000,
        "apProductivityValue": 180000,
        "processAccelerationValue": 70000,
        "totalYear1Value": 650000,
        "totalYear2Value": 850000,
        "paybackMonths": 5,
        "threeYearROI": 5.5
    }

    # Timeline
    rom_data["timeline"] = {
        "phases": [
            {"name": "Discovery & Setup", "weeks": "1-2", "description": "Kickoff, data source identification, environment setup"},
            {"name": "AI Development", "weeks": "3-5", "description": "SAP integration, AI model configuration, prompt engineering"},
            {"name": "UI & Testing", "weeks": "6-8", "description": "Power Apps interface, workflow testing, UAT"},
            {"name": "Live Pilot", "weeks": "9-12", "description": "Process 100+ invoices, measure results, prepare recommendations"}
        ],
        "milestones": [
            {"date": "February 2026", "event": "Scope & budget alignment"},
            {"date": "End of February", "event": "MSA terms complete"},
            {"date": "Early March", "event": "SOW delivered"},
            {"date": "Mid-March", "event": "Finalize agreements"},
            {"date": "April 6", "event": "Project kick-off"},
            {"date": "Late June", "event": "Pilot live"}
        ]
    }

    # Use cases
    rom_data["useCases"] = [
        {
            "id": 1,
            "name": "AI Vendor Communication Bot",
            "type": "Quick Win",
            "challenge": "Vendor inquiries take 2-3 days via portal or 5-10 days via email",
            "solution": "AI-powered chatbot handling invoice status and payment queries in real-time",
            "targets": [
                {"metric": "Deflection rate", "value": "60-70%"},
                {"metric": "Response time", "value": "< 5 minutes"},
                {"metric": "AP time savings", "value": "20-30%"}
            ]
        },
        {
            "id": 2,
            "name": "Order-to-Cash AI Automation",
            "type": "Strategic Win",
            "challenge": "$70M trapped in receivables, 60-day DSO due to ECN delays",
            "solution": "AI identifies ECNs, calculates pricing impact, generates invoice drafts",
            "targets": [
                {"metric": "Invoice cycle time", "value": "30-40% reduction"},
                {"metric": "AI accuracy", "value": "80%+ vs manual"},
                {"metric": "Working capital", "value": "Faster unlock"}
            ]
        }
    ]

    # Calculate totals
    if rom_data["phases"]:
        total_low = sum(p.get("pointsLow", 0) for p in rom_data["phases"].values())
        total_high = sum(p.get("pointsHigh", 0) for p in rom_data["phases"].values())
        total_features = sum(p.get("features", 0) for p in rom_data["phases"].values())
        rom_data["totals"] = {
            "pointsLow": total_low,
            "pointsHigh": total_high,
            "totalFeatures": total_features
        }

    return rom_data


def save_json(data, output_path):
    """Save data to JSON file."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)
    print(f"✓ Data saved to: {output_path}")


def main():
    # Determine ROM file path
    if len(sys.argv) > 1:
        rom_path = sys.argv[1]
    else:
        rom_path = find_rom_file()

    if not rom_path or not os.path.exists(rom_path):
        print("Error: Could not find ROM Excel file.")
        print("Usage: python update_rom_data.py [path_to_rom_excel]")
        sys.exit(1)

    # Extract data
    rom_data = extract_rom_data(rom_path)

    # Save to JSON
    output_path = "data/rom-data.json"
    save_json(rom_data, output_path)

    # Print summary
    print("\n" + "=" * 50)
    print("ROM Data Update Complete")
    print("=" * 50)
    print(f"Source: {rom_path}")
    print(f"Output: {output_path}")
    print(f"Last Updated: {rom_data['meta']['lastUpdated']}")
    print(f"\nKey Values:")
    print(f"  Investment: ${rom_data['summary']['investmentAvg']:,}")
    print(f"  Duration: {rom_data['summary']['weeksAvg']} weeks")
    print(f"  Year 1 Value: ${rom_data['valueProjections']['totalYear1Value']:,}")
    print(f"  Payback: {rom_data['valueProjections']['paybackMonths']} months")
    print("\n✓ Web presentation will reflect these values on next page load.")


if __name__ == "__main__":
    main()
